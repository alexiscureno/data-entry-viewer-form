import sys
from PyQt5.QtWidgets import QApplication, QPushButton, QLineEdit, QDialog, QSpinBox, QComboBox, QCheckBox, \
    QMainWindow, QAction, QTableWidget, QFileDialog, QMenuBar, QTableWidgetItem, QWidgetAction, QMessageBox, \
    QRadioButton, QAbstractItemView
from PyQt5 import uic
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import qdarktheme
import re


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        uic.loadUi('ui.ui', self)
        qdarktheme.setup_theme()
        self.file_path = None

        # Widgets

        # File menu bar
        self.new_file_bar = self.findChild(QAction, 'actionNew_File')
        self.open_file_bar = self.findChild(QAction, 'actionOpen_File')

        # LineEdit
        self.name_ln = self.findChild(QLineEdit, 'name_edit')

        # SpinBox
        self.age_box = self.findChild(QSpinBox, "spinBox")

        # ComboBox
        self.sus_box = self.findChild(QComboBox, "comboBox")
        self.sus_box.addItem('Subscribed')
        self.sus_box.addItem('Unsubscribed')
        self.sus_box.addItem('Other')

        # CheckBox
        self.employ_box = self.findChild(QCheckBox, "checkBox")

        # pushButton
        self.insert_btn = self.findChild(QPushButton, "pushButton")
        self.insert_btn.clicked.connect(self.insert_data)

        self.delete_btn = self.findChild(QPushButton, "pushButton_2")
        self.delete_btn.clicked.connect(self.delete_selected_row)

        # Radio Buttons
        self.dark_mode = self.findChild(QRadioButton, "radioButton")
        self.dark_mode.toggled.connect(self.set_dark_mode)

        self.light_mode = self.findChild(QRadioButton, "radioButton_2")
        self.light_mode.toggled.connect(self.set_light_mode)

        # Actions menu bar
        self.new_file_bar.triggered.connect(self.new_file)
        self.open_file_bar.triggered.connect(self.open_file)

        # Table Widget
        self.table_widget = self.findChild(QTableWidget, 'tableWidget')
        self.table_widget.cellDoubleClicked.connect(self.select_row)
        self.table_widget.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table_widget.clicked.connect(self.clear_widgets)

    def set_dark_mode(self):
        qdarktheme.setup_theme()
    def set_light_mode(self):
        qdarktheme.setup_theme("light")

    def insert_data(self):

        if self.file_path:
            name = self.name_ln.text()
            age = self.age_box.value()
            sub = self.sus_box.currentText()
            if self.employ_box.isChecked():
                employ = "Employed"
            else:
                employ = "Unemployed"

            if re.match(r'^[A-Za-z\s]+$', name) and name.strip() and age > 0 and sub:
                selected_indexes = self.table_widget.selectedIndexes()
                if selected_indexes:
                    # Get the first selected row index
                    row_index = selected_indexes[0].row()

                    # Update the data in the table widget
                    self.update_data_in_table(row_index, name, age, sub, employ)

                    # Update the data in the Excel file
                    self.update_data_in_excel(row_index, name, age, sub, employ)
                else:

                    self.add_data_to_excel(self.file_path, name, age, sub, employ)
                    self.add_data_to_table(name, age, sub, employ)

                self.name_ln.clear()
                self.age_box.setValue(self.age_box.minimum())
                self.sus_box.setCurrentIndex(-1)
                self.employ_box.setChecked(False)
            else:
                QMessageBox.warning(self, "Invalid Input",
                                    "Please enter a valid name (letters only) and a positive age.")
        else:
            QMessageBox.warning(self, "No File Selected", "Please create or open a file before inserting data.")
            self.clear_widgets()
    def clear_widgets(self):
        self.name_ln.clear()
        self.age_box.setValue(self.age_box.minimum())
        self.sus_box.setCurrentIndex(-1)
        self.employ_box.setChecked(False)
    def add_data_to_excel(self, file_path, name, age, sub, employ):
        # Open the existing workbook
        wb = openpyxl.load_workbook(file_path)

        # Select the active worksheet
        ws = wb.active

        # Get the last row index
        last_row = ws.max_row + 1

        # Write the data to the worksheet
        ws.cell(row=last_row, column=1, value=name)
        ws.cell(row=last_row, column=2, value=age)
        ws.cell(row=last_row, column=3, value=sub)
        ws.cell(row=last_row, column=4, value=employ)

        # Save the workbook
        wb.save(file_path)

    def update_data_in_excel(self, row_index, name, age, sub, employ):
        if self.file_path:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb.active
            ws.cell(row=row_index + 2, column=1, value=name)  # +2 to account for header row and 0-indexing
            ws.cell(row=row_index + 2, column=2, value=age)
            ws.cell(row=row_index + 2, column=3, value=sub)
            ws.cell(row=row_index + 2, column=4, value=employ)
            wb.save(self.file_path)

    def add_data_to_table(self, name, age, sub, employ):
        # Get the current row count in the table widget
        current_row_count = self.table_widget.rowCount()

        # Insert a new row in the table widget
        self.table_widget.insertRow(current_row_count)

        # Set the data in the table widget
        self.table_widget.setItem(current_row_count, 0, QTableWidgetItem(name))
        self.table_widget.setItem(current_row_count, 1, QTableWidgetItem(str(age)))
        self.table_widget.setItem(current_row_count, 2, QTableWidgetItem(sub))
        self.table_widget.setItem(current_row_count, 3, QTableWidgetItem(employ))

    def update_data_in_table(self, row_index, name, age, sub, employ):
        self.table_widget.setItem(row_index, 0, QTableWidgetItem(name))
        self.table_widget.setItem(row_index, 1, QTableWidgetItem(str(age)))
        self.table_widget.setItem(row_index, 2, QTableWidgetItem(sub))
        self.table_widget.setItem(row_index, 3, QTableWidgetItem(employ))

    def new_file(self):
        wb = Workbook()
        new_file_name = QFileDialog.getSaveFileName(
            parent=self,
            caption=' Save file location and name',
            directory='',
            filter='Data File (*.xlsx *.csv*.) ;; Excel File (*.xlsx *.xls)',
            initialFilter='Excel File (*.xlsx *.xls)'
        )
        if new_file_name[0]:
            file_path = new_file_name[0]

            # Save the workbook
            wb.save(file_path)
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            ws.cell(row=1, column=1, value="Name")
            ws.cell(row=1, column=2, value="Age")
            ws.cell(row=1, column=3, value="Subscription")
            ws.cell(row=1, column=4, value="Employment")
            wb.save(file_path)

            self.file_path = file_path

            self.load_data(self.file_path)

    def open_file(self):
        opened_file_name = QFileDialog.getOpenFileName(
            parent=self,
            caption='Open file',
            directory='',
            filter='Data File (*.xlsx *.csv*.) ;; Excel File (*.xlsx *.xls)',
            initialFilter='Excel File (*.xlsx *.xls)'
        )
        if opened_file_name[0]:
            file_path = opened_file_name[0]
            self.file_path = file_path
            wb = openpyxl.load_workbook(opened_file_name[0])

            wb.save(opened_file_name[0])

            self.load_data(self.file_path)

    def load_data(self, path_file):
        path = path_file
        df = pd.read_excel(path)

        cols = list(df.columns)
        rows = df.to_numpy().tolist()
        x = len(cols)
        y = len(rows)

        self.tableWidget.setRowCount(y)
        self.tableWidget.setColumnCount(x)

        for j in range(x):
            header = QTableWidgetItem(cols[j])
            self.tableWidget.setHorizontalHeaderItem(j, header)

            for i in range(y):
                data = str(rows[i][j])
                if data == 'nan':
                    data = ''
                self.tableWidget.setItem(i, j, QTableWidgetItem(data))

    def select_row(self, row_index):
        if row_index >= 0:
            row_data = self.get_selected_row_values(row_index)
            print("Selected Row Data:", row_data)
            if row_data:

                self.name_ln.setText(row_data[0])
                self.age_box.setValue(int(row_data[1]))
                self.sus_box.setCurrentText(row_data[2])
                self.employ_box.setChecked(row_data[3] == "Employed")
            else:
                self.clear_widgets()

    def get_selected_row_values(self, row_index):
        row_data = []
        for column in range(self.table_widget.columnCount()):
            item = self.table_widget.item(row_index, column)
            if item is not None:
                row_data.append(item.text())
        return row_data

    def delete_selected_row(self):
        selected_row_indexes = self.table_widget.selectedIndexes()
        if not selected_row_indexes:
            return

        # Get the row indexes of the selected cells
        row_indexes = set()
        for index in selected_row_indexes:
            row_indexes.add(index.row())


        # Sort the row indexes in descending order to ensure correct deletion
        row_indexes = sorted(row_indexes, reverse=True)

        # Delete the rows from the QTableWidget
        for row_index in row_indexes:
            self.table_widget.removeRow(row_index)

        # Delete the rows from the Excel file
        if self.file_path:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb.active
            for row_index in row_indexes:
                ws.delete_rows(row_index + 2)  # +2 to account for header row and 0-indexing
            wb.save(self.file_path)

    def delete_selected_rows(self):
        selected_row_indexes = self.table_widget.selectedIndexes()
        if not selected_row_indexes:
            return

        # Get the row indexes of the selected rows
        row_indexes = set()
        for index in selected_row_indexes:
            row_indexes.add(index.row())

        # Sort the row indexes in descending order to ensure correct deletion
        row_indexes = sorted(row_indexes, reverse=True)

        # Delete the rows from the QTableWidget
        for row_index in row_indexes:
            self.table_widget.removeRow(row_index)

        # Delete the rows from the Excel file
        if self.file_path:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb.active
            for row_index in row_indexes:
                ws.delete_rows(row_index + 2)  # +2 to account for header row and 0-indexing
            wb.save(self.file_path)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
